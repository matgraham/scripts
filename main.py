from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from openpyxl import load_workbook, Workbook

def parse_data_into_dict(data):
    list_of_players = []
    list_of_stats = []
    stat_dict = {
        "age": 'B', 'gp': 'C','w':'D', 'l':'E','min': 'F', 'pts':'G', 'fgm':'H','fga':'I','fg%':'J',
        "3pm":'K','3pa':'L','ftm':'M','fta':'N', 'ft%':'O','oreb':'P','dreb':'Q','reb':'R','ast':'S',
        'tov':'T','stl':'U','blk':'V','pf':'W',"dd2":'X','td3':'y'
    }
    excelfile = 'nbastats.xlsx'
    wb = load_workbookd(excelfile)
    ws = wb[wb.sheetnames[0]]
    for row in range(1, ws.max_row+1):
        for col in "A":
            cell_name = "{}{}".format(col,row)
            list_of_players.append(ws[cell_name].value.lower())
            for col in stat_dict[data]:
                cell_name = "{}{}".format(col,row)
                list_of_stats.append(ws[cell_name].value)
    return dict(zip(list_of_players, list_of_stats))

app = Flask(__name__)
@app.route('/', methods=['GET','POST'])

def send_sms():
    msg = request.form['Body'].lower()
    typomsg = 'send 1st and last names of two players followed by a stat (GP, W, L, MIN, PTS, FG%, 3P%, FT%, REB, AST, STL, BLK). Check for typos!'            
    if len(player_and_stat) == 5:
        player1 = player_and_stat[0] + '' + player_and_stat[1]
        player2 = player_and_stat[2] + '' + player_and_stat[3]
        stat = player_and_stat[4]
        player_stat_map = parse_data_into_dict(stat)
        if player1 in player_stat_map.keys() and player2 in player_stat_map.keys():
            if player_stat_map[player1] > player_stat_map[player2]:
                ret = MessagingResponse().message(player1 + " 's total " + str(player_stat_map[player1]) + ", higher than " + player2 + "\'s " + str(player_stat_map[player2]))
            else:
                ret = MessagingResponse().message(player2 + " 's total " + str(player_stat_map[player2]) + ", higher than " + player1 + "\'s " + str(player_stat_map[player1]))
        else: #check
            ret = MessagingResponse().message("check both players' names (first and last!)")
    else:
        ret = MessagingResponse().message(typomsg)
        return str(ret)

if __name__ == "__main__":
    app.run(debug=True)